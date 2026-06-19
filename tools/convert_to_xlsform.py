#!/usr/bin/env python3
"""Convert Open Scale Definition (OSD) format to XLSForm Excel (.xlsx).

Usage:
    python3 convert_to_xlsform.py <scale_directory> [--output FILE] [--lang LANG] [--all-langs]
    python3 convert_to_xlsform.py scales/grit/ --output grit.xlsx
    python3 convert_to_xlsform.py scales/GQ6/ --lang de
    python3 convert_to_xlsform.py scales/MES/  --all-langs

Reads {code}.json and {code}.{lang}.json (or {code}.osd) from a scale directory
and generates an XLSForm Excel file suitable for import into:
    KoboToolbox, SurveyCTO, ODK Collect, OpenClinica, DHIS2, and others.

Supported conversions:
  - likert           -> select_one {list_name}  (appearance: likert)
  - multi            -> select_one {list_name}
  - multicheck       -> select_multiple {list_name}
  - dropdown         -> select_one {list_name}  (appearance: minimal)
  - short            -> text
  - long             -> text  (appearance: multiline)
  - number           -> integer / decimal  (depending on validation)
  - date             -> date
  - vas              -> range  (parameters: start=min end=max step=1)
  - grid             -> begin group (appearance: field-list) + select_one per row
  - inst             -> note
  - section          -> note  (section header)
  - constant_sum     -> multiple integer fields
  - image/imageresponse -> note  (with image URL hint)
  - Scoring          -> calculate fields (XPath sum/mean expressions)
  - visible_when     -> relevant column (XPath expressions)
  - Multilingual     -> label::Language (code) columns via --all-langs

Notes:
  - Reverse coding is handled in calculate field XPath expressions.
  - Grid items are expanded into a field-list group with one select_one per row.
  - Scoring with 'scores' (composite) references other calculate fields.
  - Choices lists are deduplicated: items sharing the same response scale share
    one choices list_name.
  - XLSForm names are sanitized to [A-Za-z_][A-Za-z0-9_]* and validated.
"""

import json
import re
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SURVEY_COLUMNS = [
    "type",
    "name",
    "label",
    "hint",
    "relevant",
    "required",
    "calculation",
    "appearance",
    "constraint",
    "constraint_message",
    "default",
    "parameters",
]

CHOICES_COLUMNS = [
    "list_name",
    "name",
    "label",
]

SETTINGS_COLUMNS = [
    "form_title",
    "form_id",
    "version",
    "default_language",
]

# Unicode → ASCII equivalents for XLSForm/Enketo XML compatibility
_UNICODE_NORMALIZE = str.maketrans({
    "–": "-",    # en-dash
    "—": "--",   # em-dash
    "‘": "'",    # left single quote
    "’": "'",    # right single quote / apostrophe
    "“": '"',    # left double quote
    "”": '"',    # right double quote
    "…": "...",  # ellipsis
    "≥": ">=",   # greater-than-or-equal
    "≤": "<=",   # less-than-or-equal
    "±": "+/-",  # plus-minus
    "°": " deg", # degree sign
    "×": "x",    # multiplication sign
    "·": "*",    # middle dot / multiplication
    "−": "-",    # minus sign (not hyphen)
    "→": "->",   # right arrow
    "←": "<-",   # left arrow
    "↑": "^",    # up arrow
    "↓": "v",    # down arrow
})


def normalize_text(s):
    """Replace common non-ASCII typography with ASCII equivalents."""
    if not s:
        return s
    return s.translate(_UNICODE_NORMALIZE)


# Header fill colours (pale blue / pale green)
HEADER_FILL_SURVEY = PatternFill("solid", fgColor="DCEBF7")
HEADER_FILL_CHOICES = PatternFill("solid", fgColor="DCF7DC")
HEADER_FILL_SETTINGS = PatternFill("solid", fgColor="F7EDDC")
HEADER_FONT = Font(bold=True)

# OSD operators → XPath operators
XPATH_OP_MAP = {
    "==":          "=",
    "=":           "=",
    "equals":      "=",
    "!=":          "!=",
    "not_equals":  "!=",
    ">":           ">",
    "greater_than": ">",
    "<":           "<",
    "less_than":   "<",
    ">=":          ">=",
    "<=":          "<=",
}

# Operators that don't use a value RHS
XPATH_UNARY_OPS = {"is_answered", "is_not_answered", "is_true", "is_false"}


# ---------------------------------------------------------------------------
# Shared helpers (matching convert_to_redcap.py style)
# ---------------------------------------------------------------------------

def find_definition_file(scale_dir):
    """Find the main .json definition file."""
    p = Path(scale_dir)
    code = p.name
    definition = p / f"{code}.json"
    if definition.exists():
        return definition, code
    for f in sorted(p.glob("*.json")):
        if not re.match(r".*\.\w{2}(-\w+)?\.json$", f.name):
            return f, f.stem
    for f in sorted(p.glob("*.osd")):
        return f, f.stem
    return None, code


def load_translation(scale_dir, code, lang="en"):
    """Load a translation file."""
    p = Path(scale_dir)
    for pattern in [f"{code}.{lang}.json", f"{code}.pbl-{lang}.json"]:
        path = p / pattern
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    # Fallback: try loading from .osd file
    for osd_file in p.glob("*.osd"):
        try:
            with open(osd_file, "r", encoding="utf-8") as f:
                osd_data = json.load(f)
            translations = osd_data.get("translations", {})
            if lang in translations:
                return translations[lang]
            if translations:
                return next(iter(translations.values()))
        except (json.JSONDecodeError, KeyError):
            pass
    return {}


def load_all_translations(scale_dir, code, primary_lang="en"):
    """Return {lang: translation_dict} for all available languages.

    Looks in both .osd embedded translations and legacy separate files.
    """
    p = Path(scale_dir)
    translations = {}

    # Try .osd first
    for osd_file in sorted(p.glob("*.osd")):
        try:
            with open(osd_file, "r", encoding="utf-8") as f:
                osd_data = json.load(f)
            embedded = osd_data.get("translations", {})
            translations.update(embedded)
        except (json.JSONDecodeError, KeyError):
            pass
        break  # only one .osd expected

    # Legacy: {code}.{lang}.json and {code}.pbl-{lang}.json
    for tf in p.glob(f"{code}.*.json"):
        m = re.match(rf"{re.escape(code)}\.(\w{{2}}(?:-\w+)?)\.json$", tf.name)
        if m:
            lang_code = m.group(1)
            if lang_code not in translations:
                try:
                    with open(tf, "r", encoding="utf-8") as f:
                        translations[lang_code] = json.load(f)
                except (json.JSONDecodeError, OSError):
                    pass
    for tf in p.glob(f"{code}.pbl-*.json"):
        m = re.match(rf"{re.escape(code)}\.pbl-(\w{{2}}(?:-\w+)?)\.json$", tf.name)
        if m:
            lang_code = m.group(1)
            if lang_code not in translations:
                try:
                    with open(tf, "r", encoding="utf-8") as f:
                        translations[lang_code] = json.load(f)
                except (json.JSONDecodeError, OSError):
                    pass

    # Always ensure primary_lang is present (even if empty fallback)
    if primary_lang not in translations:
        t = load_translation(scale_dir, code, primary_lang)
        if t:
            translations[primary_lang] = t

    return translations


def get_text(translations, key, fallback=None):
    """Get translated text, trying case-insensitive lookup."""
    if not key:
        return fallback if fallback is not None else ""
    if key in translations:
        return translations[key]
    key_lo = key.lower()
    for k, v in translations.items():
        if k.lower() == key_lo:
            return v
    return fallback if fallback is not None else key


def strip_html(text):
    """Strip HTML tags and normalize typography for XLSForm compatibility."""
    if not text:
        return ""
    return normalize_text(re.sub(r"<[^>]+>", "", str(text)).strip())


def get_effective_scale(question, definition):
    """Resolve the effective response scale opts for a likert question."""
    rs_id = question.get("response_scale")
    if rs_id:
        rs = definition.get("response_scales", {}).get(rs_id)
        if rs:
            return rs
    return definition.get("likert_options", {})


def get_likert_labels(question, definition, translations):
    """Get likert labels for a question (per-item or scale-level)."""
    if "likert_labels" in question:
        return [get_text(translations, lbl) for lbl in question["likert_labels"]]
    eff = get_effective_scale(question, definition)
    if eff.get("labels"):
        return [get_text(translations, lbl) if lbl else None
                for lbl in eff["labels"]]
    points = question.get("likert_points", eff.get("points", 5))
    min_val = eff.get("min", 1)
    return [str(min_val + i) for i in range(points)]


def get_likert_min(question, definition):
    """Get the minimum numeric value for likert scoring."""
    return get_effective_scale(question, definition).get("min", 1)


def get_likert_max(question, definition):
    """Get the maximum numeric value for likert scoring."""
    eff = get_effective_scale(question, definition)
    points = question.get("likert_points", eff.get("points", 5))
    min_val = eff.get("min", 1)
    return min_val + points - 1


# ---------------------------------------------------------------------------
# XLSForm name validation / sanitisation
# ---------------------------------------------------------------------------

def make_xlsform_name(raw):
    """Convert a question ID to a valid XLSForm field name.

    XLSForm names must:
    - Start with a letter or underscore
    - Contain only letters, digits, underscores
    - Be unique within the form (caller handles uniqueness)
    """
    name = re.sub(r"[^A-Za-z0-9_]", "_", raw)
    if name and not (name[0].isalpha() or name[0] == "_"):
        name = "q_" + name
    return name


def validate_xlsform_name(name):
    """Return True if name is a valid XLSForm identifier."""
    return bool(re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name))


# ---------------------------------------------------------------------------
# visible_when → XPath relevant expression
# ---------------------------------------------------------------------------

def _simple_condition_to_xpath(cond):
    """Convert a single leaf condition dict to an XPath fragment."""
    if not isinstance(cond, dict):
        # String shorthand like "true" or "false"
        return "true()" if str(cond).lower() in ("true", "1", "yes") else "false()"
    # Parameter conditions are pre-launch settings, not runtime fields —
    # variant resolution handles them; skip in relevant expressions.
    if "parameter" in cond and "item" not in cond and "question" not in cond:
        return "true()"
    item = cond.get("item") or cond.get("question")
    if not item:
        return "true()"
    op_raw = cond.get("operator", cond.get("op", "=="))
    name = make_xlsform_name(str(item))
    ref = f"${{{name}}}"

    # Unary operators — no RHS value
    if op_raw == "is_answered":
        return f"{ref} != ''"
    if op_raw == "is_not_answered":
        return f"{ref} = ''"
    if op_raw == "is_true":
        return f"{ref} = 'true'"
    if op_raw == "is_false":
        return f"{ref} = 'false'"

    op = XPATH_OP_MAP.get(op_raw, "=")
    value = cond.get("value", "")
    # Quote strings; leave numbers bare
    try:
        float(str(value))
        val_str = str(value)
    except (TypeError, ValueError):
        val_str = f"'{value}'"
    return f"{ref} {op} {val_str}"


def visible_when_to_relevant(visible_when):
    """Recursively convert OSD visible_when to XLSForm relevant XPath."""
    if not visible_when:
        return ""
    if isinstance(visible_when, str):
        # String shorthand: bare parameter name → always true at runtime
        return "true()"
    if not isinstance(visible_when, dict):
        return "true()"
    if "all" in visible_when:
        parts = [visible_when_to_relevant(c) for c in visible_when["all"]]
        return "(" + " and ".join(parts) + ")"
    if "any" in visible_when:
        parts = [visible_when_to_relevant(c) for c in visible_when["any"]]
        return "(" + " or ".join(parts) + ")"
    if "not" in visible_when:
        inner = visible_when_to_relevant(visible_when["not"])
        return f"not({inner})"
    return _simple_condition_to_xpath(visible_when)


# ---------------------------------------------------------------------------
# Choices list management
# ---------------------------------------------------------------------------

class ChoicesRegistry:
    """Tracks unique choices lists and assigns list_name identifiers.

    Likert items sharing the same response scale share one list.
    multi/multicheck items with per-item options get their own list.
    """

    def __init__(self):
        # key → list_name  (key is a frozenset-of-tuples for likert,
        #                    or "item:{qid}" for per-item lists)
        self._key_to_name = {}
        # list_name → [(value, label_key_or_text), ...]
        self._lists = {}
        self._counter = {}

    def _unique_name(self, base):
        n = self._counter.get(base, 0) + 1
        self._counter[base] = n
        return base if n == 1 else f"{base}_{n}"

    def register_likert(self, question, definition, primary_translations,
                        all_translations=None):
        """Register a likert choices list; return list_name.

        Two questions share a list when they have the same response_scale id
        (or both use the global likert_options) AND the same per-item labels.
        """
        has_per_item_labels = "likert_labels" in question
        rs_id = question.get("response_scale", "__global__")

        if has_per_item_labels:
            # Per-item override: unique list keyed by question id
            qid = question["id"]
            key = f"item:{qid}"
        else:
            # Shared: key on response_scale id + points
            eff = get_effective_scale(question, definition)
            points = question.get("likert_points", eff.get("points", 5))
            key = f"likert:{rs_id}:{points}"

        if key in self._key_to_name:
            return self._key_to_name[key]

        # Build the list entry
        eff = get_effective_scale(question, definition)
        min_val = eff.get("min", 1)
        labels_primary = get_likert_labels(question, definition,
                                           primary_translations)
        points = len(labels_primary)

        choices = []
        for i, lbl in enumerate(labels_primary):
            val = min_val + i
            # Collect labels across all languages
            lang_labels = {}
            if all_translations:
                raw_label_keys = (question.get("likert_labels")
                                  or eff.get("labels", []))
                lbl_key = (raw_label_keys[i]
                           if i < len(raw_label_keys) else None)
                for lang, trans in all_translations.items():
                    if lbl_key:
                        lang_labels[lang] = get_text(trans, lbl_key,
                                                     str(val))
                    else:
                        lang_labels[lang] = str(val)
            choices.append({
                "name": str(val),
                "label": strip_html(lbl) if lbl else str(val),
                "lang_labels": lang_labels,
            })

        # Choose list_name
        if has_per_item_labels:
            base = make_xlsform_name(question["id"]) + "_opts"
        elif rs_id != "__global__":
            base = make_xlsform_name(rs_id) + "_opts"
        else:
            base = "likert_main"

        list_name = self._unique_name(base)
        self._key_to_name[key] = list_name
        self._lists[list_name] = choices
        return list_name

    def register_options(self, question, translations, all_translations=None,
                         list_name_hint=None):
        """Register a per-item choices list for multi/multicheck/dropdown.

        Returns list_name.
        """
        qid = question["id"]
        key = f"item:{qid}"
        if key in self._key_to_name:
            return self._key_to_name[key]

        options = question.get("options", [])
        choices = []
        for i, opt in enumerate(options):
            if isinstance(opt, dict):
                val = str(opt.get("value", i + 1))
                text_key = opt.get("text_key", opt.get("value", ""))
                label = strip_html(get_text(translations, str(text_key),
                                            str(text_key)))
                lang_labels = {}
                if all_translations:
                    for lang, trans in all_translations.items():
                        lang_labels[lang] = strip_html(
                            get_text(trans, str(text_key), label))
            else:
                val = str(i + 1)
                label = strip_html(get_text(translations, str(opt), str(opt)))
                lang_labels = {}
                if all_translations:
                    for lang, trans in all_translations.items():
                        lang_labels[lang] = strip_html(
                            get_text(trans, str(opt), label))
            choices.append({
                "name": val,
                "label": label,
                "lang_labels": lang_labels,
            })

        base = list_name_hint or (make_xlsform_name(qid) + "_opts")
        list_name = self._unique_name(base)
        self._key_to_name[key] = list_name
        self._lists[list_name] = choices
        return list_name

    def register_grid_columns(self, question, translations,
                              all_translations=None):
        """Register choices for a grid question's columns. Returns list_name."""
        qid = question["id"]
        key = f"grid_cols:{qid}"
        if key in self._key_to_name:
            return self._key_to_name[key]

        columns = question.get("columns", [])
        choices = []
        for i, col in enumerate(columns):
            if isinstance(col, dict):
                val = str(col.get("value", i + 1))
                text_key = col.get("text_key", str(val))
                label = strip_html(get_text(translations, text_key, text_key))
                lang_labels = {}
                if all_translations:
                    for lang, trans in all_translations.items():
                        lang_labels[lang] = strip_html(
                            get_text(trans, text_key, label))
            else:
                val = str(i + 1)
                label = strip_html(get_text(translations, str(col), str(col)))
                lang_labels = {}
                if all_translations:
                    for lang, trans in all_translations.items():
                        lang_labels[lang] = strip_html(
                            get_text(trans, str(col), label))
            choices.append({
                "name": val,
                "label": label,
                "lang_labels": lang_labels,
            })

        list_name = self._unique_name(make_xlsform_name(qid) + "_cols")
        self._key_to_name[key] = list_name
        self._lists[list_name] = choices
        return list_name

    def all_lists(self):
        """Yield (list_name, choices) for all registered lists."""
        for list_name, choices in self._lists.items():
            yield list_name, choices


# ---------------------------------------------------------------------------
# Scoring → XPath calculate expressions
# ---------------------------------------------------------------------------

def _scoring_items_list(score_def):
    """Return list of (item_id, coding) from a score definition.

    Handles two OSD formats:
      - 'items': ["id1", "id2", ...]  with optional 'item_coding': {"id1": -1}
      - 'items': {"id1": 1, "id2": -1, ...}  (coding embedded in dict)
    """
    items_raw = score_def.get("items", [])
    item_coding = score_def.get("item_coding", {})

    if isinstance(items_raw, dict):
        # Embedded coding
        return [(item_id, coding)
                for item_id, coding in items_raw.items()]
    else:
        # List + optional item_coding dict
        return [(item_id, item_coding.get(item_id, 1))
                for item_id in items_raw]


def build_xlsform_calculate(score_id, score_def, definition, code,
                             all_score_names, active_item_ids=None):
    """Build an XPath calculate expression for XLSForm.

    active_item_ids: set of item IDs present in the flattened form; scoring
    references to absent items (e.g. from inactive variants) are skipped.
    Returns expression string or empty string if not computable.
    """
    method = score_def.get("method", "")
    raw_pairs = _scoring_items_list(score_def)
    # Filter to items that actually exist in this variant of the form
    if active_item_ids is not None:
        item_pairs = [(iid, c) for iid, c in raw_pairs if iid in active_item_ids]
    else:
        item_pairs = raw_pairs
    scores_refs = score_def.get("scores", [])  # composite: references other scores

    # Determine reverse-sum floor for likert scale
    likert_opts = definition.get("likert_options", {})
    min_val = likert_opts.get("min", 1)
    points = likert_opts.get("points", 5)
    max_val = min_val + points - 1
    reverse_sum = min_val + max_val

    def item_expr(item_id, coding):
        name = make_xlsform_name(item_id)
        if coding == -1:
            return f"({reverse_sum} - ${{{name}}})"
        return f"${{{name}}}"

    def score_ref_expr(sid):
        """Reference to another calculate field."""
        calc_name = make_xlsform_name(f"{code}_{sid}")
        return f"${{{calc_name}}}"

    if method in ("sum_coded", "sum"):
        if scores_refs and not item_pairs:
            # Composite: sum of other score fields
            parts = [score_ref_expr(s) for s in scores_refs]
            return " + ".join(parts)
        if scores_refs and item_pairs:
            # Mixed: sum of both
            parts = [item_expr(iid, c) for iid, c in item_pairs]
            parts += [score_ref_expr(s) for s in scores_refs]
            return " + ".join(parts)
        if not item_pairs:
            return ""
        parts = [item_expr(iid, c) for iid, c in item_pairs]
        return " + ".join(parts)

    elif method == "mean_coded":
        if scores_refs and not item_pairs:
            parts = [score_ref_expr(s) for s in scores_refs]
            n = len(parts)
            return f"({' + '.join(parts)}) div {n}"
        if not item_pairs:
            return ""
        parts = [item_expr(iid, c) for iid, c in item_pairs]
        n = len(parts)
        inner = " + ".join(parts)
        if scores_refs:
            ref_parts = [score_ref_expr(s) for s in scores_refs]
            inner = " + ".join(parts + ref_parts)
            n += len(ref_parts)
        return f"({inner}) div {n}"

    elif method == "weighted_sum":
        weights = score_def.get("weights", {})
        if scores_refs:
            parts = []
            for s in scores_refs:
                w = weights.get(s, 1)
                parts.append(f"({w} * {score_ref_expr(s)})")
            return " + ".join(parts)
        if not item_pairs:
            return ""
        parts = []
        for iid, c in item_pairs:
            w = weights.get(iid, 1)
            expr = item_expr(iid, c)
            parts.append(f"({w} * {expr})")
        return " + ".join(parts)

    elif method == "sum_correct":
        # Binary accuracy sum
        if not item_pairs:
            return ""
        parts = [f"${{{make_xlsform_name(iid)}}}" for iid, _ in item_pairs]
        return " + ".join(parts)

    else:
        # Fallback: plain sum
        if scores_refs and not item_pairs:
            parts = [score_ref_expr(s) for s in scores_refs]
            return " + ".join(parts)
        if not item_pairs:
            return ""
        parts = [f"${{{make_xlsform_name(iid)}}}" for iid, _ in item_pairs]
        return " + ".join(parts)


def _apply_transform(expression, transform):
    """Wrap an XPath expression with scoring transforms (multiply, divide, etc.)."""
    if not transform or not expression:
        return expression
    expr = expression
    for step in transform:
        op = step.get("op", "")
        val = step.get("value", 1)
        if op == "multiply":
            expr = f"({expr}) * {val}"
        elif op == "divide":
            expr = f"({expr}) div {val}"
        elif op == "add":
            expr = f"({expr}) + {val}"
        elif op == "subtract":
            expr = f"({expr}) - {val}"
        elif op == "round":
            expr = f"round({expr}, {int(val)})"
    return expr


# ---------------------------------------------------------------------------
# Survey row builders
# ---------------------------------------------------------------------------

def _make_survey_row(type_="", name="", label="", hint="", relevant="",
                     required="", calculation="", appearance="",
                     constraint="", constraint_message="", default="",
                     parameters=""):
    """Return a dict with all survey columns."""
    return {
        "type": type_,
        "name": name,
        "label": label,
        "hint": hint,
        "relevant": relevant,
        "required": required,
        "calculation": calculation,
        "appearance": appearance,
        "constraint": constraint,
        "constraint_message": constraint_message,
        "default": default,
        "parameters": parameters,
    }


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_xlsform(definition, translations, all_translations=None):
    """Generate XLSForm data from OSD format.

    Args:
        definition: Parsed OSD definition dict (already flattened of variants).
        translations: Primary language translation dict {key: text}.
        all_translations: Optional {lang: {key: text}} for multilingual output.

    Returns:
        (survey_rows, choices_registry, langs)
        where survey_rows is a list of row dicts and langs is the ordered list
        of language codes for multilingual columns.
    """
    scale_info = definition.get("scale_info", {})
    code = scale_info.get("code", "scale")
    questions = definition.get("items") or definition.get("questions", [])
    scoring = definition.get("scoring", {})
    pages = definition.get("pages", None)

    # Language ordering: primary first, rest alphabetical
    langs = []
    if all_translations:
        # Sort: primary lang first (inferred as the translations dict lang),
        # then remaining alphabetically
        primary = _guess_primary_lang(translations, all_translations)
        langs = [primary] + sorted(l for l in all_translations if l != primary)
    # Single-language fallback — langs stays empty → no ::lang columns

    survey_rows = []
    choices = ChoicesRegistry()

    # Build page → first-item mapping for section notes
    q_to_page = {}
    if pages:
        for page in pages:
            for item_id in page.get("items", []):
                q_to_page[item_id] = page

    first_page_item = {}
    if pages:
        for page in pages:
            items_on_page = page.get("items", [])
            if items_on_page:
                first_page_item[items_on_page[0]] = page

    # Identify first likert item for question_head note emission
    likert_opts = definition.get("likert_options", {})
    global_question_head_key = likert_opts.get("question_head", "")
    first_likert_id = next(
        (q["id"] for q in questions if q.get("type") == "likert"),
        None
    )

    for q in questions:
        qtype = q.get("type", "")
        qid = q["id"]
        field_name = make_xlsform_name(qid)
        text_key = q.get("text_key", qid)
        label = get_text(translations, text_key, qid)
        visible_when = q.get("visible_when")
        relevant = visible_when_to_relevant(visible_when) if visible_when else ""
        is_required = q.get("required", qtype in ("likert", "vas", "multi",
                                                    "grid", "multicheck"))
        required_val = "yes" if is_required else ""

        # Build multilingual label dict
        lang_labels = {}
        if langs:
            for lang in langs:
                trans = all_translations.get(lang, {})
                lang_labels[lang] = get_text(trans, text_key, label)

        # --- Page section note ---
        if qid in first_page_item:
            page = first_page_item[qid]
            title_key = page.get("title_key", "")
            if title_key:
                page_title = get_text(translations, title_key, "")
                if page_title:
                    page_name = make_xlsform_name(
                        page.get("id", f"page_{qid}"))
                    row = _make_survey_row(
                        type_="note",
                        name=page_name,
                        label=page_title,
                    )
                    if langs:
                        for lang in langs:
                            trans = all_translations.get(lang, {})
                            row[f"label::{_lang_tag(lang)}"] = get_text(
                                trans, title_key, page_title)
                    survey_rows.append(row)

        # --- Global question_head note (emitted before first likert item) ---
        if qid == first_likert_id and global_question_head_key:
            head_text = get_text(translations, global_question_head_key, "")
            if head_text:
                head_row = _make_survey_row(
                    type_="note",
                    name=make_xlsform_name(global_question_head_key),
                    label=strip_html(head_text),
                )
                if langs:
                    for lang in langs:
                        trans = all_translations.get(lang, {})
                        head_row[f"label::{_lang_tag(lang)}"] = strip_html(
                            get_text(trans, global_question_head_key, head_text))
                survey_rows.append(head_row)

        # --- Type dispatch ---

        if qtype == "likert":
            list_name = choices.register_likert(
                q, definition, translations, all_translations if langs else None)
            row = _make_survey_row(
                type_=f"select_one {list_name}",
                name=field_name,
                label=strip_html(label),
                relevant=relevant,
                required=required_val,
                appearance="likert",
            )
            _add_lang_labels(row, lang_labels, langs)
            survey_rows.append(row)

        elif qtype == "multi":
            list_name = choices.register_options(
                q, translations, all_translations if langs else None)
            row = _make_survey_row(
                type_=f"select_one {list_name}",
                name=field_name,
                label=strip_html(label),
                relevant=relevant,
                required=required_val,
            )
            _add_lang_labels(row, lang_labels, langs)
            survey_rows.append(row)

        elif qtype == "multicheck":
            list_name = choices.register_options(
                q, translations, all_translations if langs else None)
            row = _make_survey_row(
                type_=f"select_multiple {list_name}",
                name=field_name,
                label=strip_html(label),
                relevant=relevant,
                required=required_val,
            )
            _add_lang_labels(row, lang_labels, langs)
            survey_rows.append(row)

        elif qtype == "dropdown":
            list_name = choices.register_options(
                q, translations, all_translations if langs else None)
            row = _make_survey_row(
                type_=f"select_one {list_name}",
                name=field_name,
                label=strip_html(label),
                relevant=relevant,
                required=required_val,
                appearance="minimal",
            )
            _add_lang_labels(row, lang_labels, langs)
            survey_rows.append(row)

        elif qtype == "short":
            hint = ""
            validation = q.get("validation", {})
            constraint_expr = ""
            constraint_msg = ""
            if validation:
                vmin = validation.get("min", validation.get("number_min"))
                vmax = validation.get("max", validation.get("number_max"))
                if vmin is not None and vmax is not None:
                    constraint_expr = f". >= {vmin} and . <= {vmax}"
                    constraint_msg = f"Value must be between {vmin} and {vmax}"
                elif vmin is not None:
                    constraint_expr = f". >= {vmin}"
                elif vmax is not None:
                    constraint_expr = f". <= {vmax}"
            row = _make_survey_row(
                type_="text",
                name=field_name,
                label=strip_html(label),
                relevant=relevant,
                required=required_val,
                constraint=constraint_expr,
                constraint_message=constraint_msg,
            )
            _add_lang_labels(row, lang_labels, langs)
            survey_rows.append(row)

        elif qtype == "long":
            row = _make_survey_row(
                type_="text",
                name=field_name,
                label=strip_html(label),
                relevant=relevant,
                required=required_val,
                appearance="multiline",
            )
            _add_lang_labels(row, lang_labels, langs)
            survey_rows.append(row)

        elif qtype == "number":
            val_min = q.get("min")
            val_max = q.get("max")
            # Detect decimal via validation or step
            use_decimal = q.get("decimal", False) or q.get("step") not in (
                None, 1, "1")
            xlstype = "decimal" if use_decimal else "integer"
            constraint_expr = ""
            constraint_msg = ""
            if val_min is not None and val_max is not None:
                constraint_expr = f". >= {val_min} and . <= {val_max}"
                constraint_msg = f"Value must be between {val_min} and {val_max}"
            elif val_min is not None:
                constraint_expr = f". >= {val_min}"
            elif val_max is not None:
                constraint_expr = f". <= {val_max}"
            row = _make_survey_row(
                type_=xlstype,
                name=field_name,
                label=strip_html(label),
                relevant=relevant,
                required=required_val,
                constraint=constraint_expr,
                constraint_message=constraint_msg,
            )
            _add_lang_labels(row, lang_labels, langs)
            survey_rows.append(row)

        elif qtype == "date":
            row = _make_survey_row(
                type_="date",
                name=field_name,
                label=strip_html(label),
                relevant=relevant,
                required=required_val,
            )
            _add_lang_labels(row, lang_labels, langs)
            survey_rows.append(row)

        elif qtype == "vas":
            min_val = q.get("min", q.get("min_value", 0))
            max_val = q.get("max", q.get("max_value", 100))
            step = q.get("step", 1)
            params = f"start={min_val} end={max_val} step={step}"
            hint = ""
            if "min_label" in q or "max_label" in q:
                min_lbl = get_text(translations, q.get("min_label", ""), "")
                max_lbl = get_text(translations, q.get("max_label", ""), "")
                hint_parts = []
                if min_lbl:
                    hint_parts.append(f"Low: {strip_html(min_lbl)}")
                if max_lbl:
                    hint_parts.append(f"High: {strip_html(max_lbl)}")
                hint = " / ".join(hint_parts)
            row = _make_survey_row(
                type_="range",
                name=field_name,
                label=strip_html(label),
                hint=hint,
                relevant=relevant,
                required=required_val,
                parameters=params,
            )
            _add_lang_labels(row, lang_labels, langs)
            survey_rows.append(row)

        elif qtype == "grid":
            # begin group with field-list appearance, then one select_one per row
            group_name = field_name + "_grp"
            group_label = strip_html(label)
            if q.get("question_head"):
                head_text = get_text(translations, q["question_head"], "")
                if head_text:
                    group_label = strip_html(head_text)

            list_name = choices.register_grid_columns(
                q, translations, all_translations if langs else None)

            begin_row = _make_survey_row(
                type_="begin group",
                name=group_name,
                label=group_label,
                relevant=relevant,
                appearance="field-list",
            )
            if langs:
                for lang in langs:
                    trans = all_translations.get(lang, {})
                    head_key = q.get("question_head", text_key)
                    begin_row[f"label::{_lang_tag(lang)}"] = strip_html(
                        get_text(trans, head_key, group_label))
            survey_rows.append(begin_row)

            grid_rows = q.get("rows", [])
            for j, row_key in enumerate(grid_rows):
                row_label = get_text(translations, str(row_key), str(row_key))
                row_field = f"{field_name}_{j + 1}"
                row_lang_labels = {}
                if langs:
                    for lang in langs:
                        trans = all_translations.get(lang, {})
                        row_lang_labels[lang] = get_text(
                            trans, str(row_key), row_label)
                sub_row = _make_survey_row(
                    type_=f"select_one {list_name}",
                    name=row_field,
                    label=strip_html(row_label),
                    required=required_val,
                )
                _add_lang_labels(sub_row, row_lang_labels, langs)
                survey_rows.append(sub_row)

            survey_rows.append(_make_survey_row(type_="end group",
                                                name=group_name))

        elif qtype in ("inst", "section"):
            # Sections are structural dividers — no relevant expression.
            # Skip if label is blank — pyxform requires note to have a label.
            clean_label = strip_html(label)
            if not clean_label:
                continue
            row = _make_survey_row(
                type_="note",
                name=field_name,
                label=clean_label,
                relevant=relevant if qtype == "inst" else "",
            )
            _add_lang_labels(row, lang_labels, langs)
            survey_rows.append(row)

        elif qtype == "constant_sum":
            options = q.get("options", [])
            total = q.get("total", 100)
            for i, opt in enumerate(options):
                if isinstance(opt, dict):
                    opt_text = get_text(translations,
                                        opt.get("text_key",
                                                 opt.get("value", "")))
                else:
                    opt_text = get_text(translations, str(opt), str(opt))
                opt_field = f"{field_name}_{i + 1}"
                opt_label = strip_html(opt_text)
                if i == 0:
                    opt_label = (f"{strip_html(label)} "
                                 f"[total={total}]: {opt_label}")
                opt_lang_labels = {}
                if langs:
                    for lang in langs:
                        trans = all_translations.get(lang, {})
                        if isinstance(opt, dict):
                            tkey = opt.get("text_key", opt.get("value", ""))
                            lt = get_text(trans, str(tkey), opt_text)
                        else:
                            lt = get_text(trans, str(opt), opt_text)
                        opt_lang_labels[lang] = strip_html(lt)
                constraint_expr = f". >= 0 and . <= {total}"
                sub_row = _make_survey_row(
                    type_="integer",
                    name=opt_field,
                    label=opt_label,
                    relevant=relevant if i == 0 else "",
                    required=required_val,
                    constraint=constraint_expr,
                    constraint_message=f"Enter a value between 0 and {total}",
                )
                _add_lang_labels(sub_row, opt_lang_labels, langs)
                survey_rows.append(sub_row)

        elif qtype in ("image", "imageresponse"):
            img_file = q.get("image_file", q.get("image", ""))
            hint = f"Image: {img_file}" if img_file else ""
            row = _make_survey_row(
                type_="note",
                name=field_name,
                label=strip_html(label),
                hint=hint,
                relevant=relevant,
            )
            _add_lang_labels(row, lang_labels, langs)
            survey_rows.append(row)

        else:
            # Unknown type — emit as note
            row = _make_survey_row(
                type_="note",
                name=field_name,
                label=f"[{qtype}] {strip_html(label)}",
                relevant=relevant,
            )
            _add_lang_labels(row, lang_labels, langs)
            survey_rows.append(row)

    # --- Scoring calculate fields ---
    # calculate rows must be at the top level (not inside a group) — ODK rejects
    # groups whose only children are calculate fields (invisible in body).
    if scoring:
        all_score_names = set(scoring.keys())
        active_item_ids = {
            q.get("id") for q in questions
            if isinstance(q, dict) and q.get("id")
        }

        for score_id, score_def in scoring.items():
            if not isinstance(score_def, dict):
                continue

            method = score_def.get("method", "")
            expression = build_xlsform_calculate(
                score_id, score_def, definition, code, all_score_names,
                active_item_ids=active_item_ids)
            if not expression:
                continue

            # Apply post-hoc transform if any
            transform = score_def.get("transform")
            if transform:
                expression = _apply_transform(expression, transform)

            calc_name = make_xlsform_name(f"{code}_{score_id}")
            desc = score_def.get("description", "")
            note_label = (f"{score_id} ({method})"
                          + (f" - {normalize_text(desc)}" if desc else ""))
            note_label = note_label[:200]

            survey_rows.append(_make_survey_row(
                type_="calculate",
                name=calc_name,
                label=note_label,
                calculation=expression,
            ))

    return survey_rows, choices, langs


def _guess_primary_lang(primary_translations, all_translations):
    """Identify which key in all_translations matches primary_translations."""
    if not all_translations:
        return "en"
    # Match by identity or equality
    for lang, trans in all_translations.items():
        if trans is primary_translations or trans == primary_translations:
            return lang
    # Fallback: use first key
    return next(iter(all_translations))


def _lang_tag(lang):
    """Format a language tag for XLSForm multilingual column headers.

    Uses the bare code as-is (e.g. 'en', 'de', 'fr', 'zh-CN').
    """
    return lang


def _add_lang_labels(row, lang_labels, langs):
    """Add label::{lang} columns to a survey row dict."""
    for lang in langs:
        col = f"label::{_lang_tag(lang)}"
        row[col] = lang_labels.get(lang, "")


# ---------------------------------------------------------------------------
# Excel / openpyxl writer
# ---------------------------------------------------------------------------

def _style_header_row(ws, fill):
    """Apply bold + background fill to the first row of a worksheet."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, max_col_width=80):
    """Auto-size columns up to max_col_width characters."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cv = str(cell.value or "")
                max_len = max(max_len, len(cv))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_col_width)


def write_xlsform(survey_rows, choices_registry, langs, definition, code,
                  output_path):
    """Write the three-sheet XLSForm workbook to output_path."""
    wb = openpyxl.Workbook()

    # ---------------------------------------------------------------- survey
    ws_survey = wb.active
    ws_survey.title = "survey"

    # Build dynamic column list (insert label::lang columns after label)
    survey_cols = list(SURVEY_COLUMNS)
    if langs:
        label_idx = survey_cols.index("label")
        lang_label_cols = [f"label::{_lang_tag(l)}" for l in langs]
        for i, col in enumerate(lang_label_cols):
            survey_cols.insert(label_idx + 1 + i, col)

    ws_survey.append(survey_cols)
    _style_header_row(ws_survey, HEADER_FILL_SURVEY)

    for row in survey_rows:
        ws_survey.append([row.get(col, "") for col in survey_cols])

    _auto_width(ws_survey)

    # --------------------------------------------------------------- choices
    ws_choices = wb.create_sheet("choices")

    choices_cols = list(CHOICES_COLUMNS)
    if langs:
        label_idx = choices_cols.index("label")
        lang_label_cols = [f"label::{_lang_tag(l)}" for l in langs]
        for i, col in enumerate(lang_label_cols):
            choices_cols.insert(label_idx + 1 + i, col)

    ws_choices.append(choices_cols)
    _style_header_row(ws_choices, HEADER_FILL_CHOICES)

    for list_name, choice_items in choices_registry.all_lists():
        for choice in choice_items:
            row = {
                "list_name": list_name,
                "name": choice["name"],
                "label": choice["label"],
            }
            if langs:
                for lang in langs:
                    col = f"label::{_lang_tag(lang)}"
                    row[col] = choice.get("lang_labels", {}).get(lang,
                                                                   choice["label"])
            ws_choices.append([row.get(col, "") for col in choices_cols])

    _auto_width(ws_choices)

    # --------------------------------------------------------------- settings
    ws_settings = wb.create_sheet("settings")
    ws_settings.append(SETTINGS_COLUMNS)
    _style_header_row(ws_settings, HEADER_FILL_SETTINGS)

    scale_info = definition.get("scale_info", {})
    form_title = scale_info.get("name", code)
    form_id = code.lower()
    version = date.today().strftime("%Y%m%d")
    default_language = langs[0] if langs else "English (en)"

    ws_settings.append([form_title, form_id, version, default_language])
    _auto_width(ws_settings)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Convert OSD format to XLSForm Excel (.xlsx)"
    )
    parser.add_argument("scale_dir", help="Path to scale directory")
    parser.add_argument("--output", "-o",
                        help="Output .xlsx file (default: {code}.xlsx in cwd)")
    parser.add_argument("--lang", default="en",
                        help="Primary language code (default: en)")
    parser.add_argument("--all-langs", action="store_true",
                        help="Include all available languages as multilingual columns")
    parser.add_argument("--param", action="append", metavar="KEY=VALUE",
                        help="Override a scale parameter (e.g. --param version=short); "
                             "repeat for multiple params")
    args = parser.parse_args()

    scale_dir = Path(args.scale_dir)
    if not scale_dir.exists():
        print(f"Error: '{scale_dir}' not found", file=sys.stderr)
        sys.exit(1)

    def_file, code = find_definition_file(scale_dir)
    if def_file is None:
        print(f"Error: No definition file found in '{scale_dir}'",
              file=sys.stderr)
        sys.exit(1)

    with open(def_file, "r", encoding="utf-8") as f:
        definition = json.load(f)

    # Handle .osd wrapper format
    if "definition" in definition and "osd_version" in definition:
        definition = definition["definition"]

    # Variant scales: reduce to the variant active for the chosen language
    try:
        import sys as _sys
        _sys.path.insert(0, str(Path(__file__).parent))
        from osd_variants import flatten_variant
        cli_params = {}
        for kv in (args.param or []):
            if "=" in kv:
                k, v = kv.split("=", 1)
                cli_params[k.strip()] = v.strip()
        definition = flatten_variant(definition, args.lang, cli_params or None)
    except ImportError:
        print("Warning: osd_variants.py not found; skipping variant resolution",
              file=sys.stderr)

    # Load translations
    if args.all_langs:
        all_translations = load_all_translations(scale_dir, code, args.lang)
        if not all_translations:
            print(f"Warning: No translations found for any language",
                  file=sys.stderr)
            all_translations = None
        else:
            if args.lang not in all_translations:
                print(f"Warning: No translation for primary language '{args.lang}'",
                      file=sys.stderr)
        primary_translations = (all_translations.get(args.lang, {})
                                 if all_translations else {})
    else:
        primary_translations = load_translation(scale_dir, code, args.lang)
        if not primary_translations:
            print(f"Warning: No translation file found for language '{args.lang}'",
                  file=sys.stderr)
        all_translations = None

    # Determine output path
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = Path(f"{code}.xlsx")

    survey_rows, choices_registry, langs = generate_xlsform(
        definition, primary_translations, all_translations
    )

    write_xlsform(survey_rows, choices_registry, langs, definition, code,
                  output_path)

    n_items = sum(1 for r in survey_rows
                  if r.get("type", "") not in (
                      "begin group", "end group", "calculate", "note"))
    n_choices_lists = sum(1 for _ in choices_registry.all_lists())
    n_scoring = sum(1 for r in survey_rows if r.get("type") == "calculate")

    print(f"Written: {output_path}")
    print(f"  Survey rows : {len(survey_rows)} "
          f"({n_items} question fields, {n_scoring} calculate)")
    print(f"  Choices lists: {n_choices_lists}")
    if langs:
        print(f"  Languages   : {', '.join(langs)}")


if __name__ == "__main__":
    main()
