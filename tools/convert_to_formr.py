#!/usr/bin/env python3
"""Convert Open Scale Definition (OSD) format to a formr survey spreadsheet (.xlsx).

Usage:
    python3 convert_to_formr.py <scale_directory> [--output FILE] [--lang LANG]
    python3 convert_to_formr.py scales/openscales/PHQ9/ --output PHQ9_formr.xlsx
    python3 convert_to_formr.py scales/openscales/PHQ9/ --lang de

formr (https://formr.org) is a free R-based survey platform. Surveys are defined
as Excel spreadsheets uploaded to formr.org or a self-hosted instance.

The format is loosely based on XLSform but with key differences:
  - Conditional display uses the `showif` column (R expressions, not XPath)
  - Required/optional uses the `optional` column: '!' = required, '*' = optional
  - Scoring uses `calculate` type with R expressions in the `value` column
  - Likert items use `mc_button` type with inline choice1..choiceN columns
  - A separate `choices` sheet handles reusable choice lists

Item type mapping:
  likert          -> mc_button  (inline choice columns, shared choice set via choices sheet)
  multi           -> mc         (radio, inline or choices sheet)
  multicheck      -> mc_multiple (checkboxes)
  dropdown        -> select_one  (dropdown)
  short           -> text
  long            -> textarea
  number          -> number
  date            -> date
  vas             -> range       (slider)
  grid            -> mc_button rows (one per grid row, shared columns)
  inst/section    -> note        (display-only text)
  constant_sum    -> number fields (one per option)
  scoring         -> calculate   (R sum/mean expressions)

Output: a two-sheet Excel workbook:
  - survey  : item rows
  - choices : reusable choice lists (for items with > 12 options or shared scales)
"""

import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import osd_params

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

SURVEY_COLUMNS = [
    "type",
    "name",
    "label",
    "optional",
    "showif",
    "value",
    "choice1", "choice2", "choice3", "choice4", "choice5",
    "choice6", "choice7", "choice8", "choice9", "choice10",
    "choice11", "choice12",
]

CHOICES_COLUMNS = [
    "list_name",
    "name",
    "label",
]

HEADER_FILL_SURVEY  = PatternFill("solid", fgColor="DCEBF7")
HEADER_FILL_CHOICES = PatternFill("solid", fgColor="DCF7DC")
HEADER_FONT = Font(bold=True)


# ---------------------------------------------------------------------------
# OSD loaders (same pattern as other converters)
# ---------------------------------------------------------------------------

def find_definition_file(scale_dir):
    p = Path(scale_dir)
    code = p.name
    for f in sorted(p.glob("*.osd")):
        return f, f.stem
    definition = p / f"{code}.json"
    if definition.exists():
        return definition, code
    for f in sorted(p.glob("*.json")):
        if not re.match(r".*\.\w{2}(-\w+)?\.json$", f.name):
            return f, f.stem
    return None, code


def load_translation(scale_dir, code, lang="en"):
    p = Path(scale_dir)
    for pattern in [f"{code}.{lang}.json", f"{code}.pbl-{lang}.json"]:
        path = p / pattern
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    for osd_file in p.glob("*.osd"):
        try:
            with open(osd_file, "r", encoding="utf-8") as f:
                osd_data = json.load(f)
            translations = osd_data.get("translations", {})
            if lang in translations:
                return translations[lang]
            if translations:
                return next(iter(translations.values()))
        except (json.JSONDecodeError, KeyError):
            pass
    return {}


def get_text(translations, key, fallback=None):
    if not key:
        return fallback if fallback is not None else ""
    if key in translations:
        return translations[key]
    key_lo = key.lower()
    for k, v in translations.items():
        if k.lower() == key_lo:
            return v
    return fallback if fallback is not None else key


def strip_html(text):
    if not text:
        return ""
    return re.sub(r"<[^>]+>", "", str(text)).strip()


def get_effective_scale(question, definition):
    rs_id = question.get("response_scale")
    if rs_id:
        rs = definition.get("response_scales", {}).get(rs_id)
        if rs:
            return rs
    rs = definition.get("response_scales", {})
    if "default" in rs:
        return rs["default"]
    return definition.get("likert_options", {})


def get_likert_pairs(question, definition, translations):
    """Return list of (value, label) for a likert item."""
    eff = get_effective_scale(question, definition)
    min_val = eff.get("min", 1)
    label_keys = question.get("likert_labels") or eff.get("labels", [])
    points = question.get("likert_points", eff.get("points", 5))

    pairs = []
    for i in range(points):
        val = min_val + i
        if i < len(label_keys):
            label = strip_html(get_text(translations, label_keys[i], str(val)))
        else:
            label = str(val)
        pairs.append((val, label))

    if eff.get("order") == "descending":
        pairs = list(reversed(pairs))

    return pairs


def resolve_options(item, defn):
    if "options" in item:
        return item["options"]
    option_sets = defn.get("option_sets", {})
    set_key = item.get("option_set")
    if set_key and set_key in option_sets:
        return option_sets[set_key]
    if "default" in option_sets:
        return option_sets["default"]
    return []


def make_name(raw):
    name = re.sub(r"[^A-Za-z0-9_]", "_", str(raw))
    if name and not (name[0].isalpha() or name[0] == "_"):
        name = "q_" + name
    return name


# ---------------------------------------------------------------------------
# Choices registry — tracks reusable choice lists
# ---------------------------------------------------------------------------

class ChoicesRegistry:
    def __init__(self):
        self._key_to_name = {}
        self._lists = {}
        self._counter = {}

    def _unique_name(self, base):
        n = self._counter.get(base, 0) + 1
        self._counter[base] = n
        return base if n == 1 else f"{base}_{n}"

    def register_likert(self, question, definition, translations):
        """Register a shared likert choice list; return list_name."""
        rs_id = question.get("response_scale", "__global__")
        has_per_item = "likert_labels" in question
        if has_per_item:
            key = f"item:{question['id']}"
        else:
            eff = get_effective_scale(question, definition)
            points = question.get("likert_points", eff.get("points", 5))
            key = f"likert:{rs_id}:{points}"

        if key in self._key_to_name:
            return self._key_to_name[key]

        pairs = get_likert_pairs(question, definition, translations)
        choices = [{"name": str(v), "label": lbl} for v, lbl in pairs]

        if has_per_item:
            base = make_name(question["id"]) + "_opts"
        elif rs_id != "__global__":
            base = make_name(rs_id) + "_opts"
        else:
            base = "likert_main"

        list_name = self._unique_name(base)
        self._key_to_name[key] = list_name
        self._lists[list_name] = choices
        return list_name

    def register_options(self, question, translations, defn=None):
        """Register a per-item choice list; return list_name."""
        key = f"item:{question['id']}"
        if key in self._key_to_name:
            return self._key_to_name[key]

        options = resolve_options(question, defn or {})
        choices = []
        for i, opt in enumerate(options):
            if isinstance(opt, dict):
                val = str(opt.get("value", i + 1))
                text_key = opt.get("text_key", opt.get("value", ""))
                label = strip_html(get_text(translations, str(text_key), str(text_key)))
            else:
                val = str(i + 1)
                label = strip_html(get_text(translations, str(opt), str(opt)))
            choices.append({"name": val, "label": label})

        list_name = self._unique_name(make_name(question["id"]) + "_opts")
        self._key_to_name[key] = list_name
        self._lists[list_name] = choices
        return list_name

    def register_grid_columns(self, question, translations):
        """Register choices for a grid question's columns; return list_name."""
        key = f"grid_cols:{question['id']}"
        if key in self._key_to_name:
            return self._key_to_name[key]

        columns = question.get("columns", [])
        choices = []
        for i, col in enumerate(columns):
            if isinstance(col, dict):
                val = str(col.get("value", i + 1))
                text_key = col.get("text_key", str(val))
                label = strip_html(get_text(translations, text_key, text_key))
            else:
                val = str(i + 1)
                label = strip_html(get_text(translations, str(col), str(col)))
            choices.append({"name": val, "label": label})

        list_name = self._unique_name(make_name(question["id"]) + "_cols")
        self._key_to_name[key] = list_name
        self._lists[list_name] = choices
        return list_name

    def get_choices(self, list_name):
        return self._lists.get(list_name, [])

    def all_lists(self):
        for list_name, choices in self._lists.items():
            yield list_name, choices


# ---------------------------------------------------------------------------
# Inline choice columns helper
# ---------------------------------------------------------------------------

def choices_to_inline(choices):
    """Return dict of choice1..choiceN from a list of {name, label} dicts.

    formr stores the numeric value in each choiceN cell; the label is shown
    to participants. Format: "value\nlabel" per cell (formr splits on newline).
    If the list exceeds 12 items, return None (caller must use choices sheet).
    """
    if len(choices) > 12:
        return None
    result = {}
    for i, ch in enumerate(choices, 1):
        # formr choice cell format: numeric_value, newline, display_label
        result[f"choice{i}"] = f"{ch['name']}\n{ch['label']}"
    return result


# ---------------------------------------------------------------------------
# Scoring → R expressions
# ---------------------------------------------------------------------------

def _scoring_items_list(score_def):
    items_raw = score_def.get("items", [])
    item_coding = score_def.get("item_coding", {})
    if isinstance(items_raw, dict):
        return list(items_raw.items())
    return [(iid, item_coding.get(iid, 1)) for iid in items_raw]


def build_r_calculate(score_id, score_def, definition, code,
                      active_item_ids=None):
    """Build an R expression for a formr calculate field.

    Returns an R expression string, or empty string if not computable.
    """
    method = score_def.get("method", "")
    raw_pairs = _scoring_items_list(score_def)
    if active_item_ids is not None:
        item_pairs = [(iid, c) for iid, c in raw_pairs if iid in active_item_ids]
    else:
        item_pairs = raw_pairs
    scores_refs = score_def.get("scores", [])

    eff_likert = definition.get("likert_options", {})
    min_val = eff_likert.get("min", 1)
    points = eff_likert.get("points", 5)
    max_val = min_val + points - 1
    reverse_sum = min_val + max_val

    def item_expr(iid, coding):
        nm = make_name(iid)
        if coding == -1:
            return f"({reverse_sum} - {nm})"
        return nm

    def score_ref(sid):
        return make_name(f"{code}_{sid}")

    if not item_pairs and not scores_refs:
        return ""

    all_exprs = [item_expr(iid, c) for iid, c in item_pairs] + \
                [score_ref(s) for s in scores_refs]

    if method in ("sum_coded", "sum", ""):
        return " + ".join(all_exprs)

    elif method == "mean_coded":
        n = len(all_exprs)
        return f"({' + '.join(all_exprs)}) / {n}"

    elif method == "weighted_sum":
        weights = score_def.get("weights", {})
        parts = []
        for iid, c in item_pairs:
            w = weights.get(iid, 1)
            parts.append(f"({w} * {item_expr(iid, c)})")
        for s in scores_refs:
            w = weights.get(s, 1)
            parts.append(f"({w} * {score_ref(s)})")
        return " + ".join(parts)

    else:
        return " + ".join(all_exprs)


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_formr(definition, translations, lang="en", params=None):
    """Generate formr survey rows and choices registry from OSD.

    Returns (survey_rows, choices_registry).
    Each survey_row is a dict keyed by SURVEY_COLUMNS.

    When osd_params.show_header(params) is False, the respondent-facing survey
    title row is omitted so the instrument name is not revealed.
    """
    scale_info = definition.get("scale_info", {})
    code = scale_info.get("code", "scale")
    questions = definition.get("items") or definition.get("questions", [])
    scoring = definition.get("scoring", {})

    choices = ChoicesRegistry()
    survey_rows = []

    # Survey title header (respondent-facing) — guarded by show_header.
    if osd_params.show_header(params):
        survey_title = scale_info.get("name", code)
        if survey_title:
            survey_rows.append({
                "type": "note",
                "name": make_name(f"{code}_title"),
                "label": survey_title,
                "optional": "*",
                "showif": "",
                "value": "",
            })

    active_item_ids = {q["id"] for q in questions if isinstance(q, dict) and q.get("id")}

    for q in questions:
        qtype = q.get("type", "")
        qid = q["id"]
        field_name = make_name(qid)
        text_key = q.get("text_key", qid)
        label = strip_html(get_text(translations, text_key, qid))
        showif = _visible_when_to_r(q.get("visible_when"))
        is_required = q.get("required", qtype in ("likert", "vas", "multi",
                                                    "grid", "multicheck"))
        optional_val = "!" if is_required else "*"

        def base_row(ftype, fname=None, flabel=None):
            return {
                "type": ftype,
                "name": fname or field_name,
                "label": flabel if flabel is not None else label,
                "optional": optional_val,
                "showif": showif,
                "value": "",
            }

        # --- Type dispatch ---

        if qtype == "likert":
            list_name = choices.register_likert(q, definition, translations)
            ch = choices.get_choices(list_name)
            inline = choices_to_inline(ch)
            row = base_row("mc_button")
            if inline:
                row.update(inline)
            else:
                # > 12 options: reference choices sheet via list_name in type
                row["type"] = f"mc_button {list_name}"
            survey_rows.append(row)

        elif qtype == "multi":
            list_name = choices.register_options(q, translations, defn=definition)
            ch = choices.get_choices(list_name)
            inline = choices_to_inline(ch)
            row = base_row("mc")
            if inline:
                row.update(inline)
            else:
                row["type"] = f"mc {list_name}"
            survey_rows.append(row)

        elif qtype == "multicheck":
            list_name = choices.register_options(q, translations, defn=definition)
            ch = choices.get_choices(list_name)
            inline = choices_to_inline(ch)
            row = base_row("mc_multiple")
            if inline:
                row.update(inline)
            else:
                row["type"] = f"mc_multiple {list_name}"
            survey_rows.append(row)

        elif qtype == "dropdown":
            list_name = choices.register_options(q, translations, defn=definition)
            ch = choices.get_choices(list_name)
            inline = choices_to_inline(ch)
            row = base_row("select_one")
            if inline:
                row.update(inline)
            else:
                row["type"] = f"select_one {list_name}"
            survey_rows.append(row)

        elif qtype == "short":
            survey_rows.append(base_row("text"))

        elif qtype == "long":
            survey_rows.append(base_row("textarea"))

        elif qtype == "number":
            survey_rows.append(base_row("number"))

        elif qtype == "date":
            survey_rows.append(base_row("date"))

        elif qtype == "vas":
            min_val = q.get("min", q.get("min_value", 0))
            max_val = q.get("max", q.get("max_value", 100))
            step = q.get("step", 1)
            row = base_row("range")
            # formr range uses value column for R default, not config — put params in label hint
            min_lbl = strip_html(get_text(translations, q.get("min_label", ""), ""))
            max_lbl = strip_html(get_text(translations, q.get("max_label", ""), ""))
            hint_parts = [f"(0–100 slider"]
            if min_lbl:
                hint_parts.append(f"low: {min_lbl}")
            if max_lbl:
                hint_parts.append(f"high: {max_lbl}")
            hint_parts.append(f"min={min_val}, max={max_val}, step={step})")
            row["label"] = label + " " + ", ".join(hint_parts)
            # Inline choice1/choice2 encode min/max for formr range
            row["choice1"] = str(min_val)
            row["choice2"] = str(max_val)
            survey_rows.append(row)

        elif qtype == "grid":
            # One mc_button row per grid row, sharing the column choice list
            list_name = choices.register_grid_columns(q, translations)
            ch = choices.get_choices(list_name)
            inline = choices_to_inline(ch)

            grid_rows = q.get("rows", [])
            head_text = strip_html(
                get_text(translations, q.get("question_head", ""), ""))
            if head_text:
                # Emit a note for the grid header
                note_row = {
                    "type": "note",
                    "name": field_name + "_head",
                    "label": head_text,
                    "optional": "*",
                    "showif": showif,
                    "value": "",
                }
                survey_rows.append(note_row)

            for j, row_key in enumerate(grid_rows):
                row_label = strip_html(get_text(translations, str(row_key), str(row_key)))
                row_field = f"{field_name}_{j + 1}"
                row = {
                    "type": "mc_button" if inline else f"mc_button {list_name}",
                    "name": row_field,
                    "label": row_label,
                    "optional": optional_val,
                    "showif": showif if j == 0 else "",
                    "value": "",
                }
                if inline:
                    row.update(inline)
                survey_rows.append(row)

        elif qtype in ("inst", "section"):
            if label:
                survey_rows.append({
                    "type": "note",
                    "name": field_name,
                    "label": label,
                    "optional": "*",
                    "showif": showif if qtype == "inst" else "",
                    "value": "",
                })

        elif qtype == "constant_sum":
            options = q.get("options", [])
            total = q.get("total", 100)
            for i, opt in enumerate(options):
                if isinstance(opt, dict):
                    opt_text = strip_html(get_text(
                        translations, opt.get("text_key", opt.get("value", "")), ""))
                else:
                    opt_text = strip_html(get_text(translations, str(opt), str(opt)))
                opt_field = f"{field_name}_{i + 1}"
                opt_label = opt_text
                if i == 0:
                    opt_label = f"{label} [total={total}]: {opt_label}"
                survey_rows.append({
                    "type": "number",
                    "name": opt_field,
                    "label": opt_label,
                    "optional": optional_val,
                    "showif": showif if i == 0 else "",
                    "value": "",
                })

        else:
            # Unknown type — emit as note
            if label:
                survey_rows.append({
                    "type": "note",
                    "name": field_name,
                    "label": f"[{qtype}] {label}",
                    "optional": "*",
                    "showif": showif,
                    "value": "",
                })

    # --- Scoring calculate fields ---
    if scoring:
        for score_id, score_def in scoring.items():
            if not isinstance(score_def, dict):
                continue
            expr = build_r_calculate(score_id, score_def, definition, code,
                                     active_item_ids)
            if not expr:
                continue
            calc_name = make_name(f"{code}_{score_id}")
            desc = score_def.get("description", score_id)
            survey_rows.append({
                "type": "calculate",
                "name": calc_name,
                "label": f"Score: {desc}",
                "optional": "*",
                "showif": "",
                "value": expr,
            })

    return survey_rows, choices


# ---------------------------------------------------------------------------
# visible_when → R showif expression
# ---------------------------------------------------------------------------

_R_OP_MAP = {
    "==": "==", "=": "==", "equals": "==",
    "!=": "!=", "not_equals": "!=",
    ">": ">", "greater_than": ">",
    "<": "<", "less_than": "<",
    ">=": ">=", "<=": "<=",
}


def _cond_to_r(cond):
    if not isinstance(cond, dict):
        return "TRUE"
    if "parameter" in cond and "item" not in cond and "question" not in cond:
        return "TRUE"
    item = cond.get("item") or cond.get("question")
    if not item:
        return "TRUE"
    op_raw = cond.get("operator", cond.get("op", "=="))
    name = make_name(str(item))

    if op_raw == "is_answered":
        return f"!is.na({name})"
    if op_raw == "is_not_answered":
        return f"is.na({name})"
    if op_raw == "is_true":
        return f"isTRUE({name})"
    if op_raw == "is_false":
        return f"!isTRUE({name})"

    op = _R_OP_MAP.get(op_raw, "==")
    value = cond.get("value", "")
    try:
        float(str(value))
        val_str = str(value)
    except (TypeError, ValueError):
        val_str = f'"{value}"'
    return f"{name} {op} {val_str}"


def _visible_when_to_r(visible_when):
    if not visible_when:
        return ""
    if isinstance(visible_when, str):
        return "TRUE"
    if not isinstance(visible_when, dict):
        return "TRUE"
    if "all" in visible_when:
        parts = [_visible_when_to_r(c) for c in visible_when["all"]]
        return "(" + " & ".join(parts) + ")"
    if "any" in visible_when:
        parts = [_visible_when_to_r(c) for c in visible_when["any"]]
        return "(" + " | ".join(parts) + ")"
    if "not" in visible_when:
        return f"!({_visible_when_to_r(visible_when['not'])})"
    return _cond_to_r(visible_when)


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _style_header(ws, fill):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, max_w=80):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_w)


def write_formr(survey_rows, choices_registry, definition, code, output_path):
    """Write formr spreadsheet to output_path (.xlsx)."""
    wb = openpyxl.Workbook()

    # ---- survey sheet ----
    ws = wb.active
    ws.title = "survey"

    # Determine which choice columns are actually used
    max_choice = 0
    for row in survey_rows:
        for i in range(1, 13):
            if row.get(f"choice{i}"):
                max_choice = max(max_choice, i)

    used_cols = [c for c in SURVEY_COLUMNS
                 if not c.startswith("choice") or
                 int(c[6:]) <= max(max_choice, 2)]  # always show at least choice1/2
    ws.append(used_cols)
    _style_header(ws, HEADER_FILL_SURVEY)

    for row in survey_rows:
        ws.append([row.get(col, "") for col in used_cols])

    _auto_width(ws)

    # ---- choices sheet (if any lists registered) ----
    all_lists = list(choices_registry.all_lists())
    if all_lists:
        wc = wb.create_sheet("choices")
        wc.append(CHOICES_COLUMNS)
        _style_header(wc, HEADER_FILL_CHOICES)
        for list_name, items in all_lists:
            for ch in items:
                wc.append([list_name, ch["name"], ch["label"]])
        _auto_width(wc)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Convert OSD format to formr survey spreadsheet (.xlsx)"
    )
    parser.add_argument("scale_dir", help="Path to scale directory")
    parser.add_argument("--output", "-o",
                        help="Output .xlsx file (default: {code}_formr.xlsx)")
    parser.add_argument("--lang", default="en",
                        help="Language code for labels (default: en)")
    parser.add_argument("--param", action="append", metavar="KEY=VALUE",
                        help="Override a scale parameter (repeat for multiple)")
    args = parser.parse_args()

    scale_dir = Path(args.scale_dir)
    if not scale_dir.exists():
        print(f"Error: '{scale_dir}' not found", file=sys.stderr)
        sys.exit(1)

    def_file, code = find_definition_file(scale_dir)
    if def_file is None:
        print(f"Error: No definition file found in '{scale_dir}'", file=sys.stderr)
        sys.exit(1)

    with open(def_file, "r", encoding="utf-8") as f:
        definition = json.load(f)

    if "definition" in definition and "osd_version" in definition:
        definition = definition["definition"]

    # Variant resolution
    try:
        sys.path.insert(0, str(Path(__file__).parent))
        from osd_variants import flatten_variant
        cli_params = {}
        for kv in (args.param or []):
            if "=" in kv:
                k, v = kv.split("=", 1)
                cli_params[k.strip()] = v.strip()
        definition = flatten_variant(definition, args.lang, cli_params or None)
    except ImportError:
        print("Warning: osd_variants.py not found; skipping variant resolution",
              file=sys.stderr)

    translations = load_translation(scale_dir, code, args.lang)
    if not translations:
        # Try pulling from .osd embedded translations
        for osd_file in scale_dir.glob("*.osd"):
            try:
                with open(osd_file, "r", encoding="utf-8") as f:
                    osd_data = json.load(f)
                embedded = osd_data.get("translations", {})
                translations = embedded.get(args.lang, {}) or next(iter(embedded.values()), {})
            except (json.JSONDecodeError, StopIteration):
                pass
            break
    if not translations:
        print(f"Warning: No translation found for language '{args.lang}'",
              file=sys.stderr)

    # Shared OSD parameter handling: resolve effective defaults and apply
    # [name]/{name} text substitution in place. osd_params.prepare expects a
    # {lang: {key: value}} mapping; `translations` here is a flat single-language
    # dict, so wrap it (the inner dict is mutated in place).
    params = osd_params.prepare(definition, {args.lang: translations}, args.param)

    output_path = Path(args.output) if args.output else Path(f"{code}_formr.xlsx")

    survey_rows, choices_registry = generate_formr(
        definition, translations, args.lang, params=params)

    write_formr(survey_rows, choices_registry, definition, code, output_path)

    n_items = sum(1 for r in survey_rows
                  if r.get("type", "") not in ("note", "calculate"))
    n_calc = sum(1 for r in survey_rows if r.get("type") == "calculate")
    n_choices = sum(1 for _ in choices_registry.all_lists())

    print(f"Written: {output_path}")
    print(f"  Survey rows  : {len(survey_rows)} ({n_items} question fields, {n_calc} calculate)")
    print(f"  Choices lists: {n_choices}")
    print(f"  Language     : {args.lang}")
    print()
    print("Upload to formr.org:")
    print("  1. Log in → Surveys → Create new survey")
    print("  2. Upload this .xlsx file")
    print("  3. Assign the survey to a run")


if __name__ == "__main__":
    main()
